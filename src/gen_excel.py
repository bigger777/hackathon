import time
import string
import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import ProjectedPieChart, Reference, PieChart, PieChart3D

from openpyxl.chart.series import DataPoint
from sqlalchemy import create_engine

from app import db
from models import Power_line, Line_Segment, Substation, Transformer

year = datetime.datetime.now().year

hor_cell = ["A1", "B1", "D1", "E1", "F2", "G3", "H3", "I3", "J3", "L1", "M1"]
hor_words = ["№ п.п.", "Диспетчерский\nномер ЛЭП", "Напряжение кВ",
        "Год ввода\nв эксплуатацию", "Количество\nцепей", "По трассе",
        "На 1 цепь", "По трассе", "На 1 цепь", "Техническое\nсостояние",
        "Срок службы\nЛЭП"]

gor_cell = ["C1", "G2", "I2", "K2", "F1", "A6", "A7"]
gor_words = ["Наименование\nЛЭП", "Длина всего, км", 
        "Длина в т.ч\nпо участкам, км", "Марка", "Провод", "АО «Сети»",
        "Филиал «Сети 1»"]

alphabet = list(string.ascii_uppercase)

def color_cell(color, sheet, req_str):
    req = req_str.split(":")
    HeaderFill = PatternFill(start_color=color, end_color="FFFFFF",
            fill_type="solid")
    for row in sheet[req[0]:req[1]]:
        for cell in row:
            cell.fill = HeaderFill


def fill_elem(sheet, c, elem, w=0, h=0):
    cell = sheet[c]
    cell.font = Font(size=7)
    cell.value = elem
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if ((w > 0) and (h > 0)):
        sheet.row_dimensions[int(c[1:])].width = w
        sheet.column_dimensions[c[0]].height = h

def add_text(sheet):
    # Отрисовка вертикального текста
    for i in range(len(hor_cell)):
        cell = sheet[hor_cell[i]]
        cell.font = Font(size=7)
        cell.alignment = Alignment(horizontal='center', vertical='center', 
                textRotation=90)
        cell.value = hor_words[i]
    # Отрисовка вертикального текста
    for i in range(len(gor_cell)):
        cell = sheet[gor_cell[i]]
        cell.font = Font(size=7)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.value = gor_words[i]

    fill_elem(sheet, '{}{}'.format("O", 1), "Текущий год", 100, 100)

    fill_elem(sheet, '{}{}'.format("P", 1), "КЛ", 100, 100)
    fill_elem(sheet, '{}{}'.format("Q", 1), "км", 100, 100)
    fill_elem(sheet, '{}{}'.format("R", 1), "%", 100, 100)

    fill_elem(sheet, '{}{}'.format("T", 1), "ВЛ", 100, 100)
    fill_elem(sheet, '{}{}'.format("U", 1), "км", 100, 100)
    fill_elem(sheet, '{}{}'.format("V", 1), "%", 100, 100)

def gen_headers():
    book = Workbook()
    sheet = book.active

    for letter in alphabet[:5]:
        req = "{}1:{}5".format(letter, letter)
        sheet.merge_cells(req)

    sheet.merge_cells("F1:K1")
     
    sheet.merge_cells("F2:F5")
    sheet.merge_cells("G2:H2")
    sheet.merge_cells("G3:G5")
    sheet.merge_cells("H3:H5")
    
    sheet.merge_cells("K2:K5")
    sheet.merge_cells("I3:I5")
    sheet.merge_cells("J3:J5")
    sheet.merge_cells("I2:J2")
    
    sheet.merge_cells("L1:L5")
    sheet.merge_cells("M1:M5")

    color_cell("5F9EA0", sheet, "A6:M6")
    color_cell("F0E68C", sheet, "A7:M7")
    
    for i in range(6, 9):
        req = "A{}:M{}".format(i, i)
        sheet.merge_cells(req)
    
    add_text(sheet)
    
    book.save("out_lep.xlsx")
    return sheet, book

def parse_input(path):
  
    read_book = load_workbook(path) 
    sheet = read_book.active 

    values = []

    rows = sheet.rows
    for row in rows:
        for cell in row:
            if (cell.value != None):
                print(cell.value)


def get_lines(sheet, pl):
    seg_count = 0
    lss = pl.lines_segments
    for ls in lss:
        if (ls.primary_line):
            seg_count += 1
    return seg_count

def fill_pl(sheet, pl, idx):
    fill_elem(sheet, 'A{}'.format(idx), pl.id, 100, 100)
    fill_elem(sheet, 'B{}'.format(idx), pl.control_number, 100, 100)
    fill_elem(sheet, 'D{}'.format(idx), pl.voltage_class, 100, 100)
    fill_elem(sheet, 'E{}'.format(idx), pl.year_of_commissioning, 100, 100)
    fill_elem(sheet, 'F{}'.format(idx), pl.lines_segments[0].lines_amount)

def fill_ls(sheet, ls, idx):
    fill_elem(sheet, 'B{}'.format(idx), ls.control_number, 100, 100)
    fill_elem(sheet, 'I{}'.format(idx), ls.segment_length, 100, 100)
    fill_elem(sheet, 'J{}'.format(idx), ls.segment_length * ls.lines_amount,
            300, 300)
    fill_elem(sheet, 'F{}'.format(idx), ls.lines_amount, 300, 300)
    out = ""
    for wires_mark in ls.wires_mark.split("|"):
        out += wires_mark + "\n"
    fill_elem(sheet, 'K{}'.format(idx), out, 300, 300)

def calc_proc(ar):
    s = 0
    for i in range(len(ar)):
        if (sum(ar) != 0):
            s += (100.0*ar[i]/sum(ar))
    return s

def fill_lines(book, sheet, ar, w):
    s = sum(ar)

    fill_elem(sheet, '{}{}'.format(w[0], 2), ar[2], 100, 100)
    fill_elem(sheet, '{}{}'.format(w[0], 3), ar[1], 100, 100)
    fill_elem(sheet, '{}{}'.format(w[0], 4), ar[0], 100, 100)
    fill_elem(sheet, '{}{}'.format(w[0], 5), sum(ar), 100, 100)
   
    if s == 0:
        s = 1
    fill_elem(sheet, '{}{}'.format(w[1], 2), 100.0*ar[2]/s, 100, 100)
    fill_elem(sheet, '{}{}'.format(w[1], 3), 100.0*ar[1]/s, 100, 100)
    fill_elem(sheet, '{}{}'.format(w[1], 4), 100.0*ar[0]/s, 100, 100)
    fill_elem(sheet, '{}{}'.format(w[1], 5), calc_proc(ar), 100, 100)


def gen_graph(book, sheet, kl, vl):
    # отрисовка шаблона
    fill_lines(book, sheet, kl, ["Q", "R"])
    fill_lines(book, sheet, vl, ["U", "V"])
    a = ["P", "T"]
    # заполнение шаблона
    for w in a:
        fill_elem(sheet, '{}{}'.format(w, 2), "выше 50", 100, 100)
        fill_elem(sheet, '{}{}'.format(w, 3), "от 36 до 50", 100, 100)
        fill_elem(sheet, '{}{}'.format(w, 4), "до 35", 100, 100)
        fill_elem(sheet, '{}{}'.format(w, 5), "всего", 100, 100)
    
    pie = PieChart3D()
    labels = Reference(sheet, min_col=16, min_row=2, max_row=4)
    data = Reference(sheet, min_col=17, min_row=2, max_row=4)
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = "Кабельные линии"
    sheet.add_chart(pie, "O14")
    
    pie2 = PieChart3D()
    labels = Reference(sheet, min_col=20, min_row=2, max_row=4)
    data = Reference(sheet, min_col=21, min_row=2, max_row=4)
    pie2.add_data(data)
    pie2.set_categories(labels)
    pie2.title = "Воздушные линии"
    sheet.add_chart(pie2, "U14")

def pars_types(wires):
    out = ""
    for wire in set(wires[:-1].split("|")):
        out += wire + "\n"
    return out

def add_stat(arr, v, y):
    if (y > 50):
        arr[2] += v
        return
    elif ((y > 36) and (y < 50)):
        arr[1] += v
        return
    arr[0] += v

def gen_doc_lep():
    return



def fill_out_lep_xml(path_name):
    kl = [0, 0, 0]
    vl = [0, 0, 0]
    sheet, book = gen_headers()
    pls = db.session.query(Power_line).all()
    start_idx_pl = 9
    idx_pl = 8
    for pl in pls:
        # провода 
        all_wires_mark = ""
        wires_mark_without_otp = ""
        all_wires_mark_without_otp = ""
        # протяженность и индекс без отпайки
        len_without_otp = 0;
        len_without_otp_idx = 0;
        # длина и индекс общего провода
        all_len = 0;
        all_idx = 0;
        idx_pl += 1
        lss = pl.lines_segments
        # запись статичных данных
        fill_elem(sheet, 'C{}'.format(idx_pl), pl.start_segment_name + 
            "-" + pl.end_segment_name, 350, 350)
        fill_pl(sheet, pl, idx_pl)
        all_idx = idx_pl
        # проверка на наличие отпаек
        if(len(lss) > 1):
            start_idx_pl = idx_pl
            idx_pl += 1
            len_without_otp_idx = idx_pl;
            fill_pl(sheet, pl, idx_pl)
            fill_elem(sheet, 'C{}'.format(idx_pl), "протяж без отпаек", 
                    350, 350)
        else:
            fill_elem(sheet, 'K{}'.format(all_idx), pars_types(all_wires_mark))
            continue
        for ls in lss:
            idx_pl += 1
            all_wires_mark += ls.wires_mark + "|"
            # проверка на наличие отпаек
            fill_pl(sheet, pl, idx_pl)
            fill_ls(sheet, ls, idx_pl)
            all_len += ls.segment_length
            if(ls.primary_line):
                len_without_otp += ls.segment_length
                all_wires_mark_without_otp += ls.wires_mark + "|"
                fill_elem(sheet, 'C{}'.format(idx_pl), ls.start_point + '-' +
                        ls.end_point, 300, 300)
            else:
                fill_elem(sheet, 'C{}'.format(idx_pl), ls.start_point, 70, 70)
            
            # рассчет КЛ или ВЛ
            if (ls.wires_type == "kl"):
                add_stat(kl, ls.segment_length, year - ls.year_of_commissioning)
            if (ls.wires_type == "vl"):
                add_stat(vl, ls.segment_length, year - ls.year_of_commissioning)
        
        # подсчет длины проводов без отпайки
        fill_elem(sheet, 'I{}'.format(len_without_otp_idx), len_without_otp)
        fill_elem(sheet, 'J{}'.format(len_without_otp_idx), len_without_otp * 2)
        # заполнение типов проводов
        fill_elem(sheet, 'K{}'.format(len_without_otp_idx),
                pars_types(all_wires_mark_without_otp))
        fill_elem(sheet, 'K{}'.format(all_idx), pars_types(all_wires_mark))
        # сливаем ячейки, если необходимо
        sheet.merge_cells("A{}:A{}".format(start_idx_pl, idx_pl))
        sheet.merge_cells("E{}:E{}".format(start_idx_pl, idx_pl))
        # подсчет длины всего провода
        fill_elem(sheet, 'G{}'.format(all_idx), all_len)
        fill_elem(sheet, 'H{}'.format(all_idx), all_len * 2)

        start_idx_pl = idx_pl
    
    # отрисовка графика
    fill_elem(sheet, 'O3', year)
    gen_graph(book, sheet, kl, vl)
    book.save(path_name)
    
    gen_doc_lep()
    # read data from db



def fill_out_ps_xml():
    return


fill_out_lep_xml()
#fil_out_ps_xml()
#test_graph()
#parse_input("input.xlsx")

