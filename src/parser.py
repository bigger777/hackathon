# created by [BANO4KA]player7004 at Energy Hack 3
# 24.10.20

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class Parser:
    def __init__(self, file_name: str, table_type: bool = True):
        self.file_name = file_name
        self.wordbook = load_workbook(file_name)
        # True - Лэп
        # False - ПС
        self.table_type = table_type
        # Ключевые слова для ЛЭП
        self.lep_keywords = {
            "start_point": ["Наименование", "ЛЭП"],
            "include_year": ["ввода", "ввод", "год"],
            "voltage": ["напряжение", "кВ", "U", "Uном", "U ном"],
            "technical_condition": ["техничекое", "состояние", "Заключение"],
            "name": ["РЭС", "сеть", "сети"],
            "number": ["диспетчерский", "номер", "диспетчерское"]
        }
        # Ключевые слова для сегмента
        self.segment_keywords = {
            "start_point": ["Наименование", "ЛЭП"],
            "length": ["длина", "км", "трасса", "трассе"],
            "wires_num": ["количество", "цепей", "цепи", "кол-во"],
            "wire_class": ["Марка"]
        }
        # Ключевые слова для ПС(подстанции)
        self.ps_keywords = {
            "network_name": ["РЭС", "сети", "сеть"],
            "ps_name": ["ПС", "подстанции", "подстанция", "подстанционный номер"],
            "voltage": ["класс", "напряжения", "напряжение", "тип подстанции", "U", "кВ", "Uном", "U ном"],
            "include_year": ["год ввода", "ввод", "эксплуатацию", "эксплуатация", "реконструкции"]
        }
        # Ключевые слова для трансформаторов
        self.transformer_keywords = {
            "nominal_power": ["номинальная", "мощность", "МВА", "полная"],
            "manufacture_year": ["год", "изготовления"],
            "include_year": ["год", "включения"],
            "technical_condition": ["состояние", "техническое", "тех"],
            "transformer_type": ["тип"],
            "number": ["№", "номер", "диспетчерский", "диспетчерское"]
        }
        # Заголовки по которым ищем
        self.header_tags = ["№ п", "№ П"]
        # Позиция заголовка № П/П
        self.header_pos = dict((item.title(), {
            "value": "",
            "column": 0,
            "row-end": 0,
            "row": 0
        }) for item in self.wordbook.get_sheet_names())
        # Заголовки
        self.main_headers = dict((item.title(), {
            "value": [],
            "keys": [],
            "column": [],
            "row": []
        }) for item in self.wordbook.get_sheet_names())
        # Заголовки побочных таблиц
        self.add_headers = dict((item.title(), {
            "value": [],
            "keys": [],
            "column": [],
            "row": []
        }) for item in self.wordbook.get_sheet_names())
        # Позиции плашек Сети
        self.wires_pos = dict((item.title(), {
            "value": [],
            "column": [],
            "row": []
        }) for item in self.wordbook.get_sheet_names())
        # Сслыки на позиции в базе данных
        # В виде x,y
        self.main_links_to_db = dict((item.title(), {
            "start_segment_name": (1, 1),
            "end_segment_name": (1, 1),
            "year_of_commissioning": (1, 1),
            "voltage_class": (1, 1),
            "technical_condition": (1, 1),
            "network_name": (1, 1),
            "control_number": (1, 1)
        } if self.table_type else {
            # Нужен json файл для ПС таблиц
        }) for item in self.wordbook.get_sheet_names())
        # Ссылки на позиции в базе данных для побочных таблиц
        self.add_links_to_db = dict((item.title(), {
            "start_point": (1, 1),
            "end_point": (1, 1),
            "segment_length": (1, 1),
            "lines_amount": (1, 1),
            "year_of_commissioning": (1, 1),
            "wires_mark": (1, 1),
        } if self.table_type else {
            # Нужен json файл для ПС таблиц
        }) for item in self.wordbook.get_sheet_names())

        self.final_data = {}

    # Находит заголовки с их позициями
    def find_headers(self):
        for sheet in self.wordbook:
            i = 1
            for some in sheet.columns:
                data = [cell for row in sheet.iter_rows(min_col=i, max_col=i) for cell in row if cell.value]
                found = False
                for cell in data:
                    if found:
                        found = False
                        self.header_pos[sheet.title]["row-end"] = cell.row - 1
                    for value in self.header_tags:
                        if str(cell.value).find(value) != -1:
                            self.header_pos[sheet.title]["value"] = str(cell.value)
                            self.header_pos[sheet.title]["column"] = cell.column
                            self.header_pos[sheet.title]["row"] = cell.row
                            found = True
                data = [cell for row in sheet.iter_rows(min_col=i, max_col=i, min_row=self.header_pos[sheet.title]
                ["row"], max_row=self.header_pos[sheet.title]["row-end"]) for cell in row if cell.value]
                for cell in data:
                    for key, keywords in (self.lep_keywords.items() if self.table_type else self.ps_keywords.items()):
                        for value in keywords:
                            if str(cell.value).find(value) != -1 and str(cell.value) not in self.main_headers[sheet.title]\
                                    ["value"] and key not in self.main_headers[sheet.title]["keys"]:
                                self.main_headers[sheet.title]["value"].append(str(cell.value))
                                self.main_headers[sheet.title]["keys"].append(key)
                                self.main_headers[sheet.title]["column"].append(cell.column)
                                self.main_headers[sheet.title]["row"].append(cell.row)
                    for key, keywords in (self.segment_keywords.items() if self.table_type else self.transformer_keywords.items()):
                        for value in keywords:
                            if str(cell.value).find(value) != -1 and str(cell.value) not in self.add_headers[sheet.title]\
                                    ["value"] and key not in self.add_headers[sheet.title]["keys"]:
                                self.add_headers[sheet.title]["value"].append(str(cell.value))
                                self.add_headers[sheet.title]["keys"].append(key)
                                self.add_headers[sheet.title]["column"].append(cell.column)
                                self.add_headers[sheet.title]["row"].append(cell.row)
                i += 1

    # Находит положения плашек Сети
    def find_wires(self):
        for sheet in self.wordbook:
            for some in sheet.columns:
                data = [cell for row in sheet.iter_rows(min_col=1, max_col=1) for cell in row if cell.value]
                for cell in data:
                    if str(cell.value).find("Сети") != -1 and str(cell.value) not in self.wires_pos[sheet.title]["value"]:
                        self.wires_pos[sheet.title]["value"].append(str(cell.value))
                        self.wires_pos[sheet.title]["column"].append(cell.column)
                        self.wires_pos[sheet.title]["row"].append(cell.row)

    def linker_main(self):
        for sheet in self.wordbook:
            for i in range(len(self.main_headers[sheet.title]["keys"])):
                key = self.main_headers[sheet.title]["keys"][i]
                if key == "start_point":
                    self.main_links_to_db[sheet.title]["start_segment_name"] = (self.main_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                    self.main_links_to_db[sheet.title]["end_segment_name"] = (self.main_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "voltage":
                    self.main_links_to_db[sheet.title]["voltage_class"] = (self.main_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "include_year":
                    self.main_links_to_db[sheet.title]["year_of_commissioning"] = (self.main_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "technical_condition":
                    self.main_links_to_db[sheet.title]["technical_condition"] = (
                    self.main_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "name":
                    self.main_links_to_db[sheet.title]["network_name"] = (
                    self.main_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "number":
                    self.main_links_to_db[sheet.title]["control_number"] = (
                    self.main_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])

    def linker_add(self):
        for sheet in self.wordbook:
            for i in range(len(self.add_headers[sheet.title]["keys"])):
                key = self.add_headers[sheet.title]["keys"][i]
                if key == "start_point":
                    self.add_links_to_db[sheet.title]["start_point"] = (self.add_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                    self.add_links_to_db[sheet.title]["end_point"] = (self.add_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "length":
                    self.add_links_to_db[sheet.title]["segment_length"] = (self.add_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "wires_num":
                    self.add_links_to_db[sheet.title]["lines_amount"] = (self.add_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])
                elif key == "wire_class":
                    self.add_links_to_db[sheet.title]["wires_mark"] = (self.add_headers[sheet.title]["column"][i], self.header_pos[sheet.title]["row-end"])

    def get_lep_example(self):
         return {
             "start_segment_name": "",
             "end_segment_name": "",
             "year_of_commissioning": 0,
             "voltage_class": 0,
             "technical_condition": "",
             "network_name": "",
             "control_number": "",
             "lines_segments": {}
        }

    def get_segment_example(self):
        return {
            "start_point": "",
            "end_point": "",
            "segment_length": 0,
            "lines_amount": 0,
            "year_of_commissioning": 0,
            "wires_mark": "",
            "primary_line": False
        }

    def parse(self):
        for sheet in self.wordbook:
            pass
        # TODO итерация по сетям при их налчичи
        # TODO далее внутри сетей идём по номерам ПП
        # TODO далее в столбце С заголовком Наименование Лэп находим строки с именами и сохраняем их индексы,
        #  значения в шаблон
        # TODO // удалить // выпить пива
        # TODO Итерируемся по заголовкам с номерам строк найденными ранее
        # TODO Перекидываем данные в шаблон
        # TODO Профит

    # Печатает все внутренние данные
    # Дебаг хд
    def print_all(self):
        print(self.header_pos)
        print(self.main_headers)
        print(self.add_headers)
        print(self.wires_pos)
        print(self.main_links_to_db)
        print(self.add_links_to_db)
